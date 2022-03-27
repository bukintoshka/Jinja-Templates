from flask import Flask, render_template
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def homepage():
    wb = load_workbook('report_1.xlsx')
    sheet = wb.active
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=5, max_col=4):
        for cell in row:
            print(cell.value, end="|")
        print("")
    return render_template('index.html', txt = sheet)
    
    
    